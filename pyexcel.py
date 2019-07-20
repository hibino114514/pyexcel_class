 #! /usr/bin/ python
# -*- coding: utf-8 -*-

#pythonからExcelにアクセスするClass

import sys
import os.path
import os



#excel操作
import openpyxl as px
import pprint


class PyExcel:

    def __init__(self,preddir,filename=None):
        if ".xlsx" in preddir:#入力が1つのみだった場合
            self.filename=preddir
            preddir="."
        else:
            self.filename=preddir+"/"+filename
        
        
        if preddir is None:
            print "Please imput preddir,filename (+sheetname)"
        else:
            if os.path.isfile(self.filename):
                self.book=px.load_workbook(self.filename)
            else:
                print "Excelファイル作成:",self.filename
                self.book = px.Workbook()
                self.book.save(self.filename)
                



    def Help(self):
        print "***********HELP***********"
        print ""
        print "SetSheet(sheetname)              ...シート名を入力するとシートを作成してセットする"
        print "GetSheet()                       ...setされたsheet変数をreturnする"
        print "RemoveSheet(sheetname)           ...シートを削除する"
        print "InsertSheet(row,col,value,[mode])  ...シート(row,col)に値valueを入力する。modeに1を入力すると上書き防止"
        print "GetSheetValue(row,col)           ...指定したシートのセル値をreturnする"
        print ""
        print "**************************"



    def SetSheet(self,sheetname):
        sheetnames = self.book.get_sheet_names()
        if sheetname in sheetnames:
            print "Already Exsist ["+sheetname+"]."
            self.sheet = self.book.get_sheet_by_name(sheetname)

        else:
            print "シート作成:",sheetname
            self.sheet = self.book.create_sheet(title=sheetname)
        self.book.save(self.filename)
        return self.sheet





    def GetSheet(self):
        try:
            return self.sheet
        except:
            print "Please run SetSheet([sheetname])"
        return None



    def RemoveSheet(self,sheetname):
        sheetnames = self.book.get_sheet_names()
        if sheetname in sheetnames:
            self.book.remove_sheet(self.book.get_sheet_by_name(sheetname))
            self.book.save(self.filename)
        else:
            print "not fount sheet:",sheetname
        return
    

    def InsertSheet(self,row,col,val,mode=0):
        if self.sheet.cell(row=row,column=col).value is None:
            None
        else:
            if mode==1:
                print "Already exsist data this column.("+str(row)+","+str(col)+")"
                return
        self.sheet.cell(row=row,column=col).value = val
        self.book.save(self.filename)
        return

    def GetSheetValue(self,row,col):
        return self.sheet.cell(row=row,column=col).value




if __name__ == '__main__':
		ex = PyExcel()

